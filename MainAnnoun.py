import urllib.request
import logging
import os.path
import xlsxwriter
import openpyxl
import time
from bs4 import BeautifulSoup

G_HEADERS = {
    'Host': 'caigou.chinatelecom.com.cn',
    'Cookie': 'name=value; JSESSIONID=0000xrup2flCI2y1s38ilYGRXTQ:18djc0j4k; CaiGouServiceInfo=!e7fac59kWPhciaqU9I+YAUGJNqjObChF1f47G7hyMOJcKe+llhGjTigIaglrHCX3cU0RRz1SKOiRu6w=',
    'Content-Type': 'application/x-www-form-urlencoded',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36'}

G_URL = r'https://caigou.chinatelecom.com.cn/MSS-PORTAL/announcementjoin/list.do?provinceJT=NJT'

G_DICT = {'provinceJT': 'NJT', 'paging.pageSize': '10000', 'goPageNum': '1'}

G_SHEET_HEADINGS = ["省份", "是否终止", "公告名称", "公告编码", "公告类型", "创建时间日期", "开始时间", "截止时间"]


def main():
    initAnno()


'''

解决思路：利用Python 处理抓取所有界面公告信息，写入Excel，最终在Excel中进行编辑和去重。

1、G_DICT 中 paging.pageSize 属性值可以根据实际情况设置：测试过程中，基本在1000以内，当设置超过1000 是否会出现问题，暂时未知
2、G_HEADERS 中的 Cookie 可以根据实际情况来设置
3、initAnno 方法为应用程序入口
4、产生的 Excel文件，默认为当前应用程序所在目录
5、Excel文件具备在文件末尾处持续写的操作，文件命名格式按照时间戳，可以定义在一年内写一个Excel文件，只需要 将 getFileName() return 内容改为固定名称即可
6、FORM DATA 中的 goPageNum 并未发现有特殊作用，尤其是当 paging.pageSize 设置超过 100以上的时候

'''


def initAnno():
    arr_anno = []
    data = urllib.parse.urlencode(G_DICT).encode('utf-8')
    req = urllib.request.Request(url=G_URL, data=data, headers=G_HEADERS)
    res = urllib.request.urlopen(req)
    html = res.read()
    soup = BeautifulSoup(html, 'html.parser', from_encoding='gb2312')
    v_talbe = soup.find_all('table', attrs={'class': 'table_data'}, limit=30)
    first_tr = v_talbe[0]
    all_tr = first_tr.select('tr')
    i = 0
    for lp in all_tr:
        '''
         定义二维数组
        '''
        arr_anno.append([])
        '''
        table 第一行 是表头，应跳过
        '''
        if None != lp.get('class'):
            continue
        all_td = lp.select('td')
        for pd in all_td:
            value = replace_trip(pd.get_text())

            arr_anno[i].append(value)
        i += 1

    # BEGIN **************************** 开始写入Excel
    # 1、文件名
    file = get_file_name()
    # 2、检查Excel文件不存在则创建
    init_excel(file)
    # 3、写入Excel
    write_excel(arr_anno, file)
    # END **************************** 开始写入Excel


def c_print(all_anno):
    for x in all_anno:
        for y in x:
            print(y)
        print('\n')


def write_excel(arr_anno, file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    # 获得行数
    row_num = sheet.max_row
    # 获得列数
    # col_num = sheet.max_column

    current_num = row_num
    for arr in arr_anno:
        current_num += 1
        zoos = 1
        for con in arr:
            sheet.cell(current_num, zoos).value = con
            zoos += 1
    workbook.save(file)


def replace_trip(source_str):
    return source_str.replace('\n', '').replace('\r', '').lstrip()


def init_excel(file):
    if os.path.isfile(file):
        logging.error(file + ' File exists')
        return 0
    else:
        logging.warning('The ' + file + ' \t does not exist! Start creating file!')
        workbook = xlsxwriter.Workbook(file)
        # 创建一个sheet
        worksheet = workbook.add_worksheet('公告信息')
        bold = workbook.add_format({'bold': 1})
        # 写入表头
        worksheet.write_row('A1', G_SHEET_HEADINGS, bold)
        workbook.close()
        return 1


def get_file_name():
    ticks = time.strftime("%Y%m%d%H", time.localtime())
    return r'中国电信招投标公告信息-' + ticks + '.xlsx'


if __name__ == '__main__':
    main()
