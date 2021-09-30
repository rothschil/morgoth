import urllib.request
import logging
import os.path
import xlwt
import win32com.client as win32
import openpyxl
import time
from bs4 import BeautifulSoup

# Setting Headers
G_HEADERS = {
    'Host': 'caigou.chinatelecom.com.cn',
    'Cookie': 'name=value; CaiGouServiceInfo=!dMNM/vuDiA8BZNqU9I+YAUGJNqjObIaVUgLPo4CWoivbznHIwJF+VRngK7LpUps4b/5pfCDWNGQBirM=; JSESSIONID=00006WCsfp4wddDiBj1WupXLQz4:18djc0j4k',
    'Content-Type': 'application/x-www-form-urlencoded',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36'}

# 交互URL链接
G_URL = r'https://caigou.chinatelecom.com.cn/MSS-PORTAL/announcementjoin/list.do?provinceJT=NJT'

# Setting FORM DATA
G_DICT = {'provinceJT': 'NJT', 'paging.pageSize': '10', 'goPageNum': '1'}

# 设置表格头信息
G_CARS = ["省份", "是否终止", "公告名称", "公告编码", "公告类型", "创建时间日期", "开始时间", "截止时间"]


def main():
    initAnno()


'''

解决思路：利用Python 处理抓取所有界面公告信息，写入Excel，最终在Excel中进行编辑和去重。

1、G_DICT 中 paging.pageSize 属性值可以根据实际情况设置：测试过程中，基本在1000以内，当设置超过1000 是否会出现问题，暂时未知
2、G_HEADERS 中的 Cookie 可以根据实际情况来设置
3、initAnno 方法为应用程序入口
4、产生的 Excel文件，默认为当前应用程序所在目录
5、Excel文件具备在文件末尾处持续写的操作，文件命名格式按照时间戳，可以定义在一年内写一个Excel文件，只需要 将 getFileName() return 内容改为固定名称即可
6、FORM DATA 中的 goPageNum 并未发现有特殊作用，尤其是当 paging.pageSize 设置超过 100以上的时候

'''


def initAnno():
    arry_anno = []
    data = urllib.parse.urlencode(G_DICT).encode('utf-8')
    req = urllib.request.Request(url=G_URL, data=data, headers=G_HEADERS)
    res = urllib.request.urlopen(req)
    html = res.read()
    soup = BeautifulSoup(html, 'html.parser', from_encoding='gb2312')
    v_talbe = soup.find_all('table', attrs={'class': 'table_data'}, limit=30)
    first_tr = v_talbe[0]
    all_tr = first_tr.select('tr')
    i = 0
    for lp in all_tr:
        '''
         定义二维数组
        '''
        arry_anno.append([])
        '''
        table 第一行 是表头，应跳过
        '''
        if '' != lp.get('class'):
            continue
        all_td = lp.select('td')
        for pd in all_td:
            arry_anno[i].append(replace_trip(pd.get_text()))
        i += 1

    # BEGIN **************************** 开始写入Excel
    # 1、文件名
    file = get_file_name()
    # 2、检查Excel文件不存在则创建
    lag = make_excel(file)
    if lag:
        # 2.1、Excel格式转换
        xlsx_file = change(file)
    else:
        xlsx_file = file + "x"
    # 3、写入Excel
    write_excel(arry_anno, xlsx_file)
    # END **************************** 开始写入Excel


def c_print(all_anno):
    for x in all_anno:
        for y in x:
            print(y)
        print('\n')


def write_excel(all_anno, file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    # 获得行数
    row_num = sheet.max_row
    # 获得列数
    col_num = sheet.max_column

    current_num = row_num
    for arry in all_anno:
        current_num += 1
        zoos = 1
        for con in arry:
            sheet.cell(current_num, zoos).value = con
            zoos += 1
    workbook.save(file)


def replace_trip(source_str):
    return source_str.replace('\n', '').replace('\r', '').lstrip()


def make_excel(file):
    temp = file + "x"
    if os.path.isfile(temp):
        logging.error(temp + ' File exists')
        return 0
    else:
        logging.warning('The ' + file + ' does not exist! Start creating file!')
        # 创建一个workbook 设置编码
        workbook = xlwt.Workbook(encoding='utf-8')
        # 创建一个worksheet
        worksheet = workbook.add_sheet('公告信息')
        # 边框
        borders = xlwt.Borders()
        borders.left = xlwt.Borders.THIN
        borders.right = xlwt.Borders.THIN
        borders.top = xlwt.Borders.THIN
        borders.bottom = xlwt.Borders.THIN
        # 边框颜色
        borders.left_colour = 0x40
        borders.right_colour = 0x40
        borders.top_colour = 0x40
        borders.bottom_colour = 0x40

        # 表格样式
        style = xlwt.XFStyle()

        # 表格字体
        font = xlwt.Font()
        font.name = 'Times New Roman'
        font.bold = True
        font.underline = True
        style.font = font

        for i in range(len(G_CARS)):
            # 表格Header 写入excel
            worksheet.write(0, i, G_CARS[i], style)
        workbook.save(file)
        return 1


def get_file_name():
    ticks = time.strftime("%Y%m%d%H", time.localtime())
    return r'中国电信招投标公告信息-' + ticks + '.xls'


def change(file):
    # 文件绝对路径
    filename = os.path.realpath(file)
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(filename)
    # 转成xlsx格式，路径在原路径下
    file_xlsx = filename + "x"
    wb.SaveAs(file_xlsx, FileFormat=51)
    wb.Close()
    excel.Application.Quit()
    os.remove(file)
    return file_xlsx


if __name__ == '__main__':
    main()
